import os
import pickle
import numpy as np
import pandas as pd
from datetime import datetime
from flask import Flask, render_template, Response, request, jsonify, send_file, session, redirect, url_for
from functools import wraps
from flask_cors import CORS
from dotenv import load_dotenv
from supabase import create_client, Client

# ML Libraries for real-time video stream processing
import cv2
import face_recognition
import mediapipe as mp
import time

# --- Configuration and Initialization ---
load_dotenv()
app = Flask(__name__)
CORS(app) # Allow cross-origin requests, useful for development

# Session & security configuration
app.secret_key = os.getenv("FLASK_SECRET_KEY", "attendance-secret-key")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax'
)

# Load Supabase Credentials
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
ENCODINGS_FILE = os.getenv("ENCODINGS_FILE")

# Supabase Client
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Admin credentials (use environment variables in production)
ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "amardighe16@gmail.com").lower()
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "hxhn zzax rkxg qzvw")

# Global variables for ML Model
KNOWN_FACE_ENCODINGS = []
KNOWN_FACE_IDS = [] # Format: 'RollNo_Name'

# --- Attendance Session Management ---
class AttendanceSession:
    def __init__(self):
        self.detected_students = {} # {roll_no: {'name': name, 'time': timestamp}}

    def add(self, roll_no, name):
        if roll_no not in self.detected_students:
            self.detected_students[roll_no] = {
                'name': name,
                'time': datetime.now().strftime('%H:%M:%S')
            }
            print(f"✓ Added to session: {name} (Roll {roll_no})")
    
    def get_all(self):
        return self.detected_students
    
    def reset(self):
        self.detected_students = {}
        
    def count(self):
        return len(self.detected_students)

# Global instance
attendance_session = AttendanceSession()

# ==============================================================================
# SUPER-FAST BATCH PROCESSING - OPTIMIZED FOR SPEED + ACCURACY
# ==============================================================================

# Recognition settings - STRICT for accuracy
FACE_RECOGNITION_TOLERANCE = 0.48  # Stricter tolerance (lower = stricter)
CONFIDENCE_THRESHOLD = 0.54  # Higher threshold = more accurate
MIN_DETECTION_COUNT = 1  # INSTANT marking - no delay
DETECTION_COUNTER = {}  # Track consecutive detections: {RollNo: count}

# SPEED OPTIMIZATION SETTINGS
USE_CNN_MODEL = False  # HOG is much faster
FRAME_RESIZE_SCALE = 0.6  # 60% scale - balance of speed + face detail
PROCESS_EVERY_N_FRAMES = 1  # Process every frame
NUM_JITTERS = 1  # 1 jitter = fast encoding
BATCH_ENCODING = True  # Enable batch face encoding
MAX_FACES_PER_FRAME = 30  # Maximum faces to process per frame

# Pre-compute numpy arrays for vectorized matching (MUCH faster)
KNOWN_ENCODINGS_ARRAY = None  # Will be set after model load

# --- HYBRID MODEL: MediaPipe for Detection (FAST) + Dlib for Encoding (ACCURATE) ---
USE_HYBRID_MODEL = True  # Set to True for maximum speed (~0.3s per frame)

# Initialize MediaPipe Face Detection (Blazeface - extremely fast)
# Note: On headless servers (like EC2), MediaPipe may have GPU context issues
# We increase detection confidence to reduce false positives
try:
    if not hasattr(mp, 'solutions'):
        raise ImportError("MediaPipe solutions not available")
        
    mp_face_detection = mp.solutions.face_detection
    # Use model_selection=1 for FULL-RANGE detection (better for classroom with 20-25 students)
    mp_face_detector = mp_face_detection.FaceDetection(
        model_selection=1,  # 1 = full-range (5m) - better for classrooms
        min_detection_confidence=0.6  # Slightly lower to catch more faces at distance
    )
    print("MediaPipe face detection initialized (full-range mode, confidence=0.6)")
except Exception as e:
    print(f"WARNING: MediaPipe initialization failed: {e}")
    print("Falling back to pure Face_Recognition (dlib) model. Performance may be slower.")
    USE_HYBRID_MODEL = False
    mp_face_detector = None

# Minimum face size (as fraction of frame) to filter out tiny false detections
MIN_FACE_SIZE_RATIO = 0.05  # Face must be at least 5% of frame width


def is_authenticated() -> bool:
    """Returns True when the current session is logged in."""
    return session.get('user_email') is not None


def authenticate_user(email: str, password: str) -> bool:
    """Validates credentials against the configured admin account."""
    if not email or not password:
        return False
    return email.strip().lower() == ADMIN_EMAIL and password.strip() == ADMIN_PASSWORD


def login_required(view_func):
    """Decorator to protect routes that require authentication."""

    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if is_authenticated():
            return view_func(*args, **kwargs)

        # If the request expects JSON (e.g., fetch), return 401 instead of redirecting
        if request.is_json or request.accept_mimetypes.best == 'application/json':
            return jsonify({"error": "Authentication required"}), 401

        next_url = request.url if request.method == 'GET' else request.referrer
        return redirect(url_for('login', next=next_url))

    return wrapper

# Load the face recognition model
def load_model():
    """Loads the pre-trained face encodings from the pickle file."""
    global KNOWN_FACE_ENCODINGS, KNOWN_FACE_IDS, KNOWN_ENCODINGS_ARRAY
    if not os.path.exists(ENCODINGS_FILE):
        print(f"CRITICAL: Model file '{ENCODINGS_FILE}' not found. Run train_model.py first!")
        return False
    
    with open(ENCODINGS_FILE, 'rb') as f:
        data = pickle.load(f)
        KNOWN_FACE_ENCODINGS = data["encodings"]
        KNOWN_FACE_IDS = data["names"]
    
    # PRE-COMPUTE numpy array for VECTORIZED matching (HUGE speed boost)
    # This allows batch distance calculation instead of per-face loops
    KNOWN_ENCODINGS_ARRAY = np.array(KNOWN_FACE_ENCODINGS)
    
    print(f"")
    print(f"{'='*60}")
    print(f"ML Model loaded - OPTIMIZED FOR BATCH PROCESSING")
    print(f"{'='*60}")
    print(f"Total encodings: {len(KNOWN_FACE_IDS)}")
    print(f"Unique students: {len(set([n.split('_')[0] for n in KNOWN_FACE_IDS]))}")
    print(f"Frame scale: {FRAME_RESIZE_SCALE} (50% = 2x faster)")
    print(f"Jitters: {NUM_JITTERS} (1 = 2x faster than 2)")
    print(f"Max faces/frame: {MAX_FACES_PER_FRAME}")
    print(f"Tolerance: {FACE_RECOGNITION_TOLERANCE}, Threshold: {CONFIDENCE_THRESHOLD}")
    print(f"{'='*60}")
    print(f"")
    return True

# --- Face Recognition Core Logic (Generator Function for Video Stream) ---

def get_face_locations_mediapipe(rgb_frame):
    """
    Uses MediaPipe BlazeFace for ultra-fast face detection (~15-30ms).
    Returns face locations in dlib format: (top, right, bottom, left)
    Includes filtering to reject false positives (too small, wrong aspect ratio).
    """
    height, width = rgb_frame.shape[:2]
    min_face_size = int(width * MIN_FACE_SIZE_RATIO)  # Minimum face width in pixels
    
    results = mp_face_detector.process(rgb_frame)
    face_locations = []
    
    if results.detections:
        for detection in results.detections:
            # Check detection score
            score = detection.score[0] if detection.score else 0
            if score < 0.7:  # Extra confidence check
                continue
                
            bbox = detection.location_data.relative_bounding_box
            
            # Convert relative coordinates to absolute pixels
            x = int(bbox.xmin * width)
            y = int(bbox.ymin * height)
            w = int(bbox.width * width)
            h = int(bbox.height * height)
            
            # Filter out tiny detections (likely false positives)
            if w < min_face_size or h < min_face_size:
                continue
            
            # Filter by aspect ratio (faces are roughly square, 0.6-1.8 ratio)
            aspect_ratio = w / h if h > 0 else 0
            if aspect_ratio < 0.5 or aspect_ratio > 2.0:
                continue
            
            # Clamp to image boundaries
            top = max(0, y)
            right = min(width, x + w)
            bottom = min(height, y + h)
            left = max(0, x)
            
            # Format: (top, right, bottom, left) - same as dlib
            face_locations.append((top, right, bottom, left))
    
    return face_locations


def batch_recognize_faces(face_encodings):
    """
    OPTIMIZED: Vectorized batch face recognition for 20-25 faces.
    Uses numpy broadcasting for ~10x faster matching than per-face loops.
    
    Returns: List of (name, roll_no, confidence, is_recognized) tuples
    """
    global KNOWN_ENCODINGS_ARRAY
    
    if len(face_encodings) == 0 or KNOWN_ENCODINGS_ARRAY is None or len(KNOWN_ENCODINGS_ARRAY) == 0:
        return [(\"Unknown\", None, 0.0, False) for _ in face_encodings]
    
    # Convert to numpy array for vectorized operations
    unknown_encodings = np.array(face_encodings)
    
    # VECTORIZED distance calculation: (num_unknown, num_known)
    # This is the KEY optimization - calculate ALL distances at once
    # Uses efficient numpy broadcasting instead of Python loops
    distances = np.linalg.norm(KNOWN_ENCODINGS_ARRAY - unknown_encodings[:, np.newaxis], axis=2)
    
    results = []
    for i, face_distances in enumerate(distances):
        best_match_index = np.argmin(face_distances)
        best_distance = face_distances[best_match_index]
        confidence = 1 - best_distance
        
        # Check if match passes tolerance AND confidence threshold
        is_match = best_distance <= FACE_RECOGNITION_TOLERANCE and confidence >= CONFIDENCE_THRESHOLD
        
        # Additional check: ensure match is significantly better than alternatives
        sorted_distances = np.sort(face_distances)
        if len(sorted_distances) > 1:
            # Best match should be at least 0.04 better than second best
            is_ambiguous = sorted_distances[0] + 0.04 > sorted_distances[1]
            if is_ambiguous:
                is_match = False
        
        if is_match:
            full_id = KNOWN_FACE_IDS[best_match_index]
            parts = full_id.split('_', 1)
            roll_no = parts[0]
            name = parts[1].replace('_', ' ') if len(parts) > 1 else full_id
            results.append((name, roll_no, confidence, True))
        else:
            results.append((\"Unknown\", None, confidence, False))
    
    return results


def generate_frames():
    """Captures video, detects faces, and generates the video frame stream with improved accuracy."""
    camera = cv2.VideoCapture(0) # 0 means default webcam
    if not camera.isOpened():
        print("Error: Could not open camera.")
        return

    # Set camera properties for better quality
    camera.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    camera.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    camera.set(cv2.CAP_PROP_FPS, 30)
    
    # Optimize JPEG compression for faster streaming
    camera.set(cv2.CAP_PROP_BUFFERSIZE, 1)

    frame_count = 0
    
    # Log which detection model is being used
    if USE_HYBRID_MODEL:
        print("🚀 Using HYBRID model: MediaPipe (detection) + Dlib (encoding) - FAST MODE")
    else:
        detection_model = 'cnn' if USE_CNN_MODEL else 'hog'
        print(f"Using face detection model: {detection_model}")

    while True:
        success, frame = camera.read()
        if not success:
            break
        
        frame_count += 1
        
        # Process frames at regular intervals for speed
        if frame_count % PROCESS_EVERY_N_FRAMES == 0:
            start_time = time.time()
            
            # Resize frame for faster processing
            small_frame = cv2.resize(frame, (0, 0), fx=FRAME_RESIZE_SCALE, fy=FRAME_RESIZE_SCALE)
            
            # Convert the image from BGR color (OpenCV) to RGB color (face_recognition)
            rgb_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
            
            # --- DETECTION STEP ---
            if USE_HYBRID_MODEL:
                # FAST: Use MediaPipe for detection (~15-30ms)
                face_locations = get_face_locations_mediapipe(rgb_frame)
            else:
                # SLOW: Use dlib HOG/CNN for detection (~100-500ms)
                detection_model = 'cnn' if USE_CNN_MODEL else 'hog'
                face_locations = face_recognition.face_locations(rgb_frame, model=detection_model)
            
            # --- ENCODING STEP (Always use dlib - it's accurate) ---
            face_encodings = face_recognition.face_encodings(rgb_frame, face_locations, num_jitters=1)
            
            # Calculate processing time
            processing_time = time.time() - start_time

            detected_in_frame = set()
            
            # Calculate scale factor for face locations
            scale_factor = int(1 / FRAME_RESIZE_SCALE)
            
            for face_encoding, (top, right, bottom, left) in zip(face_encodings, face_locations):
                # Scale back up face locations since frame was scaled down
                top *= scale_factor
                right *= scale_factor
                bottom *= scale_factor
                left *= scale_factor
                
                # Check if the face matches any known face with custom tolerance
                matches = face_recognition.compare_faces(
                    KNOWN_FACE_ENCODINGS, 
                    face_encoding,
                    tolerance=FACE_RECOGNITION_TOLERANCE
                )
                name = "Unknown"
                confidence = 0.0
                roll_no = None

                # Use the known face with the smallest distance to the new face
                face_distances = face_recognition.face_distance(KNOWN_FACE_ENCODINGS, face_encoding)
                best_match_index = np.argmin(face_distances)
                best_distance = face_distances[best_match_index]
                
                # Calculate confidence (1 - distance)
                confidence = 1 - best_distance
                
                # Only accept if match is found AND confidence is above threshold
                if matches[best_match_index] and confidence >= CONFIDENCE_THRESHOLD:
                    # Extract the RollNo and Name
                    full_id = KNOWN_FACE_IDS[best_match_index]
                    roll_no, student_name = full_id.split('_', 1)
                    name = student_name.replace('_', ' ')
                    
                    detected_in_frame.add(roll_no)
                    
                    # Increment detection counter for consecutive detection
                    if roll_no not in DETECTION_COUNTER:
                        DETECTION_COUNTER[roll_no] = 0
                    DETECTION_COUNTER[roll_no] += 1
                    
                    # Mark attendance only after MIN_DETECTION_COUNT consecutive detections
                    # Check if student is already in the session to avoid redundant processing
                    if roll_no not in attendance_session.get_all() and DETECTION_COUNTER[roll_no] >= MIN_DETECTION_COUNT:
                        attendance_session.add(roll_no, name)
                        print(f"✓ Attendance marked for: {name} (Roll {roll_no}) - Confidence: {confidence:.2%}")
                    
                    # Display with confidence
                    display_name = f"{name} ({confidence:.0%})"
                    color = (0, 255, 0)  # Green for recognized
                else:
                    display_name = f"Unknown ({confidence:.0%})"
                    color = (0, 0, 255)  # Red for unknown
                
                # Draw box and label on the frame
                cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
                cv2.rectangle(frame, (left, bottom - 35), (right, bottom), color, cv2.FILLED)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, display_name, (left + 6, bottom - 6), font, 0.5, (255, 255, 255), 1)
            
            # Reset counter for students not detected in this frame
            for roll in list(DETECTION_COUNTER.keys()):
                if roll not in detected_in_frame:
                    DETECTION_COUNTER[roll] = 0
            
            # Show processing speed on frame
            speed_text = f"Speed: {processing_time*1000:.0f}ms | Faces: {len(face_locations)}"
            cv2.putText(frame, speed_text, (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 0), 1)

        # Add status text to frame
        status_text = f"Detected: {attendance_session.count()} students"
        cv2.putText(frame, status_text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

        # Encode the frame as a JPEG image with higher compression for faster streaming
        ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 70])
        frame = buffer.tobytes()

        # Yield the frame in response stream
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

    camera.release()

# --- API Endpoints ---

@app.route('/')
def index():
    """Renders the main page (Dashboard/Home)."""
    return render_template('index.html')

@app.route('/student_search')
def student_search():
    """Renders the student search and dashboard page."""
    return render_template('student_search.html')

@app.route('/live_attendance')
def live_attendance():
    """Renders the live camera feed page (server-side camera - for local use only)."""
    return render_template('live_attendance.html')


@app.route('/mobile_attendance')
def mobile_attendance():
    """Renders the mobile attendance page (client-side camera - for EC2/production)."""
    return render_template('mobile_attendance.html')


@app.route('/video_feed')
def video_feed():
    """Endpoint for the streaming video feed (server-side camera)."""
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route('/api/process_frame', methods=['POST'])
def process_frame():
    """
    OPTIMIZED: Process a single frame with batch face recognition.
    Handles 20-25 faces efficiently using vectorized numpy operations.
    """
    import base64
    
    try:
        data = request.get_json()
        if not data or 'frame' not in data:
            return jsonify({"error": "No frame data provided"}), 400
        
        # Decode base64 image
        frame_data = data['frame']
        if ',' in frame_data:
            frame_data = frame_data.split(',')[1]
        
        img_bytes = base64.b64decode(frame_data)
        nparr = np.frombuffer(img_bytes, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if frame is None:
            return jsonify({"error": "Invalid image data"}), 400
        
        # ============ OPTIMIZED PROCESSING ============
        start_time = time.time()
        
        # Step 1: Resize (50% = 4x fewer pixels to process)
        small_frame = cv2.resize(frame, (0, 0), fx=FRAME_RESIZE_SCALE, fy=FRAME_RESIZE_SCALE)
        rgb_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        
        detection_start = time.time()
        
        # Step 2: Fast face detection with MediaPipe
        face_locations = []
        if USE_HYBRID_MODEL and mp_face_detector is not None:
            try:
                face_locations = get_face_locations_mediapipe(rgb_frame)
            except Exception as e:
                print(f"MediaPipe error: {e}")
        
        # Fallback to HOG if needed
        if len(face_locations) == 0:
            face_locations = face_recognition.face_locations(rgb_frame, model='hog')
        
        # Limit faces to prevent overload
        if len(face_locations) > MAX_FACES_PER_FRAME:
            face_locations = face_locations[:MAX_FACES_PER_FRAME]
        
        detection_time = time.time() - detection_start
        
        # Step 3: BATCH face encoding (single call for all faces)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations, num_jitters=NUM_JITTERS)
        
        # Step 4: BATCH recognition using vectorized numpy
        recognition_results = batch_recognize_faces(face_encodings)
        
        total_time = time.time() - start_time
        
        # Scale factor for returning coordinates
        scale_factor = 1.0 / FRAME_RESIZE_SCALE
        
        faces = []
        detected_in_frame = set()
        
        for i, ((top, right, bottom, left), (name, roll_no, confidence, recognized)) in enumerate(zip(face_locations, recognition_results)):
            # Scale back to original frame size
            top = int(top * scale_factor)
            right = int(right * scale_factor)
            bottom = int(bottom * scale_factor)
            left = int(left * scale_factor)
            
            if recognized and roll_no:
                detected_in_frame.add(roll_no)
                
                # Track consecutive detections
                if roll_no not in DETECTION_COUNTER:
                    DETECTION_COUNTER[roll_no] = 0
                DETECTION_COUNTER[roll_no] += 1
                
                # Mark attendance after sufficient detections
                if roll_no not in attendance_session.get_all() and DETECTION_COUNTER[roll_no] >= MIN_DETECTION_COUNT:
                    attendance_session.add(roll_no, name)
                    print(f"✓ MARKED: {name} (Roll {roll_no}) - {confidence:.0%}", flush=True)
            
            faces.append({
                "top": top,
                "right": right,
                "bottom": bottom,
                "left": left,
                "name": name,
                "confidence": confidence,
                "recognized": recognized,
                "roll_no": roll_no
            })
        
        # Quick cleanup of detection counter
        for roll in list(DETECTION_COUNTER.keys()):
            if roll not in detected_in_frame:
                del DETECTION_COUNTER[roll]  # Immediate cleanup
        
        return jsonify({
            "success": True,
            "processing_time_ms": round(total_time * 1000),
            "faces_detected": len(faces),
            "faces": faces,
            "session_count": attendance_session.count(),
            "students": [
                {"roll_no": k, "name": v['name'], "time": v['time']} 
                for k, v in attendance_session.get_all().items()
            ]
        })
        
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/session/status', methods=['GET'])
def session_status():
    """Returns the current list of detected students in the active session."""
    return jsonify({
        "count": attendance_session.count(),
        "students": [
            {"roll_no": k, "name": v['name'], "time": v['time']} 
            for k, v in attendance_session.get_all().items()
        ]
    })


@app.route('/api/debug/model', methods=['GET'])
def debug_model():
    """Debug endpoint to check if the model is loaded correctly."""
    unique_students = set(KNOWN_FACE_IDS)
    return jsonify({
        "model_loaded": len(KNOWN_FACE_ENCODINGS) > 0,
        "total_encodings": len(KNOWN_FACE_ENCODINGS),
        "unique_students": len(unique_students),
        "student_list": sorted(list(unique_students)),
        "tolerance": FACE_RECOGNITION_TOLERANCE,
        "confidence_threshold": CONFIDENCE_THRESHOLD,
        "frame_resize_scale": FRAME_RESIZE_SCALE,
        "hybrid_model": USE_HYBRID_MODEL
    })


@app.route('/api/session/reset', methods=['POST'])
def reset_session():
    """Resets the current attendance session."""
    global DETECTION_COUNTER
    
    # Clear the attendance session
    attendance_session.reset()
    
    # Clear detection counters
    DETECTION_COUNTER = {}
    
    print("[SESSION] Session reset - all detected students cleared", flush=True)
    
    return jsonify({
        "success": True,
        "message": "Session reset successfully",
        "count": 0,
        "students": []
    })


@app.route('/finalize_attendance', methods=['POST'])
def finalize_attendance():
    """
    Saves detected students as 'Present' in Supabase.
    Students not detected will be marked 'Absent' after cross-referencing the roster.
    Automatically detects the lecture number for today.
    """
    global DETECTION_COUNTER
    
    current_date = datetime.now().strftime('%Y-%m-%d')
    current_time = datetime.now().strftime('%H:%M:%S')

    # 1. Determine the next lecture number for today
    try:
        # Query to find the highest lecture number for today
        existing_lectures = supabase.table('attendance').select('lecture_number').eq('date', current_date).execute()
        
        if existing_lectures.data:
            # Get the maximum lecture number
            max_lecture = max([record['lecture_number'] for record in existing_lectures.data])
            lecture_number = max_lecture + 1
        else:
            # First lecture of the day
            lecture_number = 1
            
        print(f"Creating attendance for Lecture {lecture_number} on {current_date}")
        
    except Exception as e:
        print(f"Error determining lecture number: {e}")
        return jsonify({"error": f"Could not determine lecture number: {str(e)}"}), 500

    # 2. Fetch the full student roster from Supabase
    try:
        response = supabase.table('students').select('*').execute()
        full_roster = response.data
        
        if not full_roster:
            return jsonify({"error": "No students found in the database. Please add students first."}), 400
            
    except Exception as e:
        print(f"Error fetching roster: {e}")
        return jsonify({"error": f"Could not fetch student roster from Supabase: {str(e)}"}), 500

    attendance_data_to_insert = []
    
    # 3. Iterate through the full roster and determine status
    detected_students = attendance_session.get_all()
    
    # DEBUG: Print what we have in the session
    print(f"\n=== FINALIZE DEBUG ===")
    print(f"Students in session: {len(detected_students)}")
    for roll, data in detected_students.items():
        print(f"  - Roll '{roll}' ({type(roll).__name__}): {data['name']}")
    print(f"Students in DB roster: {len(full_roster)}")
    for student in full_roster[:5]:  # Show first 5
        print(f"  - Roll '{student['roll_no']}' ({type(student['roll_no']).__name__}): {student['name']}")
    print(f"======================\n")
    
    present_count = 0
    for student in full_roster:
        roll_no = str(student['roll_no'])
        name = student['name']
        
        status = "Absent"
        if roll_no in detected_students:
            status = "Present"
            present_count += 1
            print(f"✓ MATCHED: Roll {roll_no} - {name}")
        
        attendance_data_to_insert.append({
            'roll_no': int(roll_no),
            'date': current_date,
            'time': current_time,
            'lecture_number': lecture_number,
            'status': status
        })

    # 4. Insert all records into the attendance table
    try:
        response = supabase.table('attendance').insert(attendance_data_to_insert).execute()
        
        # Reset the session attendance tracker and detection counter
        attendance_session.reset()
        DETECTION_COUNTER.clear()
        
        return jsonify({
            "message": f"Attendance for Lecture {lecture_number} finalized and saved successfully.",
            "lecture_number": lecture_number,
            "date": current_date,
            "present_count": len([d for d in attendance_data_to_insert if d['status'] == 'Present']),
            "absent_count": len([d for d in attendance_data_to_insert if d['status'] == 'Absent']),
        }), 200

    except Exception as e:
        print(f"Error inserting attendance data: {e}")
        return jsonify({"error": f"Failed to save attendance: {str(e)}"}), 500


@app.route('/download_attendance', methods=['GET'])
def download_attendance():
    """Generates and downloads today's attendance report in Excel format with lecture numbers."""
    current_date = datetime.now().strftime('%Y-%m-%d')

    # Fetch today's records from Supabase
    try:
        response = supabase.table('attendance').select('date, lecture_number, status, roll_no, students(name)').eq('date', current_date).order('roll_no', desc=False).order('lecture_number', desc=False).execute()
        
        if not response.data:
            return jsonify({"error": "No attendance records found for today."}), 404
        
        # Flatten the data and create date+lecture column
        data = []
        for row in response.data:
            # Create a combined column like "26-10-2025 (L1)"
            date_lecture = f"{row['date']} (L{row['lecture_number']})"
            data.append({
                'RollNo': row['roll_no'],
                'Name': row['students']['name'] if row.get('students') else 'N/A',
                'DateLecture': date_lecture,
                'Status': row['status']
            })

        df = pd.DataFrame(data)
        
        # Create pivot table: Rows = Students, Columns = Date+Lecture, Values = Status
        pivot_df = df.pivot_table(
            index=['RollNo', 'Name'], 
            columns='DateLecture', 
            values='Status', 
            aggfunc='first'
        ).reset_index()
        
        # Sort by RollNo
        pivot_df = pivot_df.sort_values('RollNo')
        
        # Generate filename
        filename = f"attendance_report_{current_date}.xlsx"
        
        # Create Excel file with formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            pivot_df.to_excel(writer, sheet_name='Attendance', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Attendance']
            
            # Format the header row
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            # Style header row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add borders and alignment to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color code the status cells
            present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light red
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Color code based on status
                    if cell.value == "Present":
                        cell.fill = present_fill
                        cell.font = Font(bold=True, color="006100")
                    elif cell.value == "Absent":
                        cell.fill = absent_fill
                        cell.font = Font(bold=True, color="9C0006")
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 10  # RollNo
            worksheet.column_dimensions['B'].width = 20  # Name
            for col in range(3, worksheet.max_column + 1):
                worksheet.column_dimensions[worksheet.cell(1, col).column_letter].width = 18  # Date+Lecture columns
        
        # Send the file
        return send_file(filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"Error generating report: {e}")
        return jsonify({"error": f"Failed to generate download report: {str(e)}"}), 500


@app.route('/download_total_attendance', methods=['GET'])
def download_total_attendance():
    """Generates and downloads complete attendance report for all students (all dates) in Excel format."""
    try:
        # Fetch ALL attendance records from Supabase
        response = supabase.table('attendance').select('date, lecture_number, status, roll_no, students(name)').order('roll_no', desc=False).order('date', desc=False).order('lecture_number', desc=False).execute()
        
        if not response.data:
            return jsonify({"error": "No attendance records found in the database."}), 404
        
        # Flatten the data and create date+lecture column
        data = []
        for row in response.data:
            # Create a combined column like "26-10-2025 (L1)"
            date_lecture = f"{row['date']} (L{row['lecture_number']})"
            data.append({
                'RollNo': row['roll_no'],
                'Name': row['students']['name'] if row.get('students') else 'N/A',
                'DateLecture': date_lecture,
                'Status': row['status']
            })

        df = pd.DataFrame(data)
        
        # Create pivot table: Rows = Students, Columns = Date+Lecture, Values = Status
        pivot_df = df.pivot_table(
            index=['RollNo', 'Name'], 
            columns='DateLecture', 
            values='Status', 
            aggfunc='first'
        ).reset_index()
        
        # Sort by RollNo
        pivot_df = pivot_df.sort_values('RollNo')
        
        # Generate filename with current timestamp
        current_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"total_attendance_report_{current_timestamp}.xlsx"
        
        # Create Excel file with formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            pivot_df.to_excel(writer, sheet_name='Attendance', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Attendance']
            
            # Format the header row
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            # Style header row
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add borders and alignment to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color code the status cells
            present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light red
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Color code based on status
                    if cell.value == "Present":
                        cell.fill = present_fill
                        cell.font = Font(bold=True, color="006100")
                    elif cell.value == "Absent":
                        cell.fill = absent_fill
                        cell.font = Font(bold=True, color="9C0006")
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 10  # RollNo
            worksheet.column_dimensions['B'].width = 20  # Name
            for col in range(3, worksheet.max_column + 1):
                worksheet.column_dimensions[worksheet.cell(1, col).column_letter].width = 18  # Date+Lecture columns
        
        # Send the file
        return send_file(filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"Error generating total attendance report: {e}")
        return jsonify({"error": f"Failed to generate total attendance report: {str(e)}"})

    except Exception as e:
        print(f"Error generating total attendance report: {e}")
        return jsonify({"error": f"Failed to generate total attendance report: {str(e)}"}), 500


@app.route('/download_student_attendance/<int:roll_no>', methods=['GET'])
def download_student_attendance(roll_no):
    """
    Downloads attendance report for a specific student in Excel format.
    Includes lecture-wise, day-wise, and overall statistics.
    """
    try:
        # Get student info
        student_response = supabase.table('students').select('name').eq('roll_no', roll_no).single().execute()
        student_name = student_response.data['name']
        
        # Get all attendance records for this student
        attendance_response = supabase.table('attendance').select('date, time, status, lecture_number').eq('roll_no', roll_no).order('date, lecture_number', desc=False).execute()
        attendance_data = attendance_response.data
        
        if not attendance_data:
            return jsonify({"error": "No attendance records found for this student"}), 404
        
        # Create DataFrame
        df = pd.DataFrame(attendance_data)
        df['date'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m-%d')
        
        # Calculate statistics
        total_records = len(df)
        present_records = len(df[df['status'] == 'Present'])
        absent_records = len(df[df['status'] == 'Absent'])
        overall_percentage = round((present_records / total_records) * 100, 2) if total_records > 0 else 0
        
        # Calculate lecture-wise stats
        total_lectures = len(df)
        attended_lectures = present_records
        lecture_percentage = round((attended_lectures / total_lectures) * 100, 2) if total_lectures > 0 else 0
        
        # Calculate day-wise stats
        df_sorted = df.sort_values(['date', 'lecture_number'])
        first_lectures = df_sorted.groupby('date').first().reset_index()
        present_days = len(first_lectures[first_lectures['status'] == 'Present'])
        total_days = len(first_lectures)
        day_percentage = round((present_days / total_days) * 100, 2) if total_days > 0 else 0
        
        # Create Excel file with structure: lectures as rows, dates as columns
        output_filename = f"Attendance_{student_name.replace(' ', '_')}_Roll{roll_no}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(os.getcwd(), output_filename)
        
        # Prepare data for horizontal structure (dates as columns, lectures as rows)
        df_sorted = df.sort_values(['date', 'lecture_number'])
        
        # Get unique dates and lectures
        unique_dates = sorted(df_sorted['date'].unique())
        all_lecture_numbers = sorted(df_sorted['lecture_number'].unique())
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Sheet"
        
        # Styling
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        present_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        present_font = Font(color='000000', bold=True)
        
        absent_fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
        absent_font = Font(color='FFFFFF', bold=True)
        
        light_blue_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 18
        
        # Row 1: Headers - RollNo | Student Name | Date1 | Date2 | ...
        ws['A1'] = 'RollNo'
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].alignment = center_alignment
        ws['A1'].border = thin_border
        
        ws['B1'] = student_name
        ws['B1'].fill = header_fill
        ws['B1'].font = header_font
        ws['B1'].alignment = center_alignment
        ws['B1'].border = thin_border
        
        # Write date headers starting from column C
        for col_idx, date in enumerate(unique_dates, start=3):
            col_letter = get_column_letter(col_idx)
            ws[f'{col_letter}1'] = date
            ws[f'{col_letter}1'].fill = header_fill
            ws[f'{col_letter}1'].font = header_font
            ws[f'{col_letter}1'].alignment = center_alignment
            ws[f'{col_letter}1'].border = thin_border
            ws.column_dimensions[col_letter].width = 12
        
        # Rows 2+: Each row represents a lecture
        # Column A: Roll number, Column B: Lecture label, Columns C+: Status for each date
        for row_idx, lecture_num in enumerate(all_lecture_numbers, start=2):
            # Column A: Roll number
            ws[f'A{row_idx}'] = roll_no
            ws[f'A{row_idx}'].alignment = center_alignment
            ws[f'A{row_idx}'].border = thin_border
            ws[f'A{row_idx}'].fill = light_blue_fill
            
            # Column B: Lecture label
            ws[f'B{row_idx}'] = f"Lecture {lecture_num}"
            ws[f'B{row_idx}'].alignment = left_alignment
            ws[f'B{row_idx}'].border = thin_border
            ws[f'B{row_idx}'].fill = light_blue_fill
            
            # Columns C+: Status for each date
            for col_idx, date in enumerate(unique_dates, start=3):
                col_letter = get_column_letter(col_idx)
                
                # Find the status for this lecture on this date
                matching_records = df_sorted[
                    (df_sorted['date'] == date) & 
                    (df_sorted['lecture_number'] == lecture_num)
                ]
                
                if not matching_records.empty:
                    status = matching_records.iloc[0]['status']
                    ws[f'{col_letter}{row_idx}'] = status
                    ws[f'{col_letter}{row_idx}'].alignment = center_alignment
                    ws[f'{col_letter}{row_idx}'].border = thin_border
                    
                    # Color code based on status
                    if status == 'Present':
                        ws[f'{col_letter}{row_idx}'].fill = present_fill
                        ws[f'{col_letter}{row_idx}'].font = present_font
                    else:
                        ws[f'{col_letter}{row_idx}'].fill = absent_fill
                        ws[f'{col_letter}{row_idx}'].font = absent_font
                else:
                    # No record for this lecture on this date
                    ws[f'{col_letter}{row_idx}'] = ''
                    ws[f'{col_letter}{row_idx}'].border = thin_border
        
        # Save workbook
        wb.save(output_path)
        
        return send_file(output_path, as_attachment=True, download_name=output_filename)
    
    except Exception as e:
        print(f"Error generating student attendance report: {e}")
        return jsonify({"error": f"Failed to generate student report: {str(e)}"}), 500


# --- Student Dashboard API (Search and Analytics) ---

@app.route('/api/search_student', methods=['GET'])
def search_student():
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify([])

    # Use Supabase full-text search or simple filtering (Supabase handles this well)
    try:
        # Search by roll_no (exact match) or name (case-insensitive contains)
        # Supabase filtering example:
        response_roll = supabase.table('students').select('roll_no, name').eq('roll_no', query).limit(1).execute()
        if response_roll.data:
            return jsonify(response_roll.data)

        # Fallback to name search (requires PostgREST function or good RLS setup)
        # For simplicity, we assume we get a reasonable number of students and filter locally if needed, 
        # but the best practice is using Supabase's ILIKE or FTS.
        response_name = supabase.table('students').select('roll_no, name').execute() # Fetches all, inefficient but guaranteed to work
        results = [
            s for s in response_name.data 
            if query.lower() in s['name'].lower()
        ]
        return jsonify(results[:10]) # Limit to top 10 results

    except Exception as e:
        print(f"Search error: {e}")
        return jsonify({"error": "Failed to search student"}), 500


@app.route('/api/student_dashboard/<int:roll_no>', methods=['GET'])
def student_dashboard(roll_no):
    """
    Provides attendance analytics for a single student including:
    - Overall lecture-wise attendance
    - Day-wise attendance (present if attended first lecture of the day)
    - Monthly breakdown
    """
    today = datetime.now()
    
    try:
        # Step 1: Get Student Name
        student_response = supabase.table('students').select('name').eq('roll_no', roll_no).single().execute()
        student_name = student_response.data['name']
        
        # Step 2: Get ALL Attendance Records for the student (with lecture numbers)
        attendance_response = supabase.table('attendance').select('date, status, lecture_number').eq('roll_no', roll_no).order('date, lecture_number', desc=False).execute()
        attendance_data = attendance_response.data
        
        # Step 3: Get TOTAL lectures conducted (all students, all dates)
        total_lectures_response = supabase.table('attendance').select('date, lecture_number').execute()
        all_lectures = total_lectures_response.data
        
        if not attendance_data:
             return jsonify({
                "student_name": student_name,
                "roll_no": roll_no,
                "lecture_wise_stats": {
                    "attended_lectures": 0,
                    "total_lectures": len(set([(r['date'], r['lecture_number']) for r in all_lectures])) if all_lectures else 0,
                    "percentage": 0
                },
                "day_wise_stats": {
                    "present_days": 0,
                    "total_days": 0,
                    "percentage": 0
                },
                "overall_stats": {
                    "total_days": 0,
                    "present_days": 0,
                    "percentage": 0
                },
                "monthly_stats": [],
                "message": "No historical attendance data available."
            }), 200

        df_att = pd.DataFrame(attendance_data)
        df_att['date'] = pd.to_datetime(df_att['date'])
        
        # === LECTURE-WISE ATTENDANCE ===
        # Count total unique lectures conducted (across all students)
        total_unique_lectures = len(set([(r['date'], r['lecture_number']) for r in all_lectures]))
        
        # Count lectures attended by this student (status = Present)
        attended_lectures = len(df_att[df_att['status'] == 'Present'])
        
        lecture_wise_percentage = round((attended_lectures / total_unique_lectures) * 100, 2) if total_unique_lectures > 0 else 0
        
        # === DAY-WISE ATTENDANCE ===
        # Rule: Student is present for the day if they attended the FIRST lecture (lecture_number = 1)
        df_att_sorted = df_att.sort_values(['date', 'lecture_number'])
        
        # Get first lecture of each day for this student
        first_lectures = df_att_sorted.groupby('date').first().reset_index()
        
        # Count days where first lecture was attended
        present_days = len(first_lectures[first_lectures['status'] == 'Present'])
        
        # Total unique days with lectures (from all students)
        total_unique_days = len(set([r['date'] for r in all_lectures]))
        
        day_wise_percentage = round((present_days / total_unique_days) * 100, 2) if total_unique_days > 0 else 0
        
        # === OVERALL STATS (Legacy - based on all attendance records) ===
        def calculate_stats(df):
            total_records = len(df)
            present_records = len(df[df['status'] == 'Present'])
            percentage = round((present_records / total_records) * 100, 2) if total_records > 0 else 0
            return total_records, present_records, percentage

        overall_total, overall_present, overall_percentage = calculate_stats(df_att)
        
        # === MONTHLY STATS ===
        df_att['month_year'] = df_att['date'].dt.to_period('M')
        monthly_groups = df_att.groupby('month_year').apply(calculate_stats).reset_index()
        monthly_stats = []
        for index, row in monthly_groups.iterrows():
            total, present, percent = row[0]
            monthly_stats.append({
                "month": str(row['month_year']),
                "total_days": total,
                "present_days": present,
                "percentage": percent
            })
            
        return jsonify({
            "roll_no": roll_no,
            "student_name": student_name,
            "lecture_wise_stats": {
                "attended_lectures": attended_lectures,
                "total_lectures": total_unique_lectures,
                "percentage": lecture_wise_percentage
            },
            "day_wise_stats": {
                "present_days": present_days,
                "total_days": total_unique_days,
                "percentage": day_wise_percentage
            },
            "overall_stats": {
                "total_days": overall_total,
                "present_days": overall_present,
                "percentage": overall_percentage
            },
            "monthly_stats": monthly_stats
        }), 200

    except Exception as e:
        print(f"Dashboard error: {e}")
        return jsonify({"error": f"Failed to retrieve student dashboard data: {e}"}), 500


if __name__ == '__main__':
    if load_model():
        # For production deployment (Render, Railway, etc.)
        # Set debug=False and host='0.0.0.0'
        port = int(os.environ.get('PORT', 5000))
        app.run(host='0.0.0.0', port=port, debug=False)
else:
    # When running with gunicorn/wsgi, load model at import time
    load_model()
